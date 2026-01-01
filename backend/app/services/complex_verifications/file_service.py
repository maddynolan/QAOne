"""
File Download Verification Service for ArisTrace QA Platform

Features:
- Verify downloaded files exist and have correct properties
- Parse and verify CSV files
- Parse and verify Excel files  
- Parse and verify JSON/XML files
- Verify image properties (dimensions, format)
"""

import os
import re
import json
import logging
import tempfile
from datetime import datetime
from typing import Optional, List, Dict, Any, Union
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class FileInfo:
    """Basic file information"""
    path: str
    name: str
    extension: str
    size_bytes: int
    mime_type: Optional[str] = None
    created_at: Optional[datetime] = None
    modified_at: Optional[datetime] = None


@dataclass
class CSVData:
    """Parsed CSV file data"""
    headers: List[str]
    rows: List[List[str]]
    row_count: int
    column_count: int
    
    def get_cell(self, row: int, col: Union[int, str]) -> Optional[str]:
        """Get cell value by row index and column index or name"""
        if row >= len(self.rows):
            return None
        if isinstance(col, str):
            if col not in self.headers:
                return None
            col = self.headers.index(col)
        if col >= len(self.rows[row]):
            return None
        return self.rows[row][col]
    
    def get_column(self, col: Union[int, str]) -> List[str]:
        """Get all values in a column"""
        if isinstance(col, str):
            if col not in self.headers:
                return []
            col = self.headers.index(col)
        return [row[col] if col < len(row) else '' for row in self.rows]


@dataclass
class ExcelData:
    """Parsed Excel file data"""
    sheets: Dict[str, CSVData]  # Sheet name -> data
    sheet_names: List[str]


@dataclass
class ImageInfo:
    """Image file information"""
    width: int
    height: int
    format: str
    mode: str  # RGB, RGBA, etc.
    has_transparency: bool


@dataclass
class FileAssertion:
    """File assertion configuration"""
    type: str  # file_exists, size_min, size_max, csv_row_count, etc.
    expected: str
    row: Optional[int] = None
    col: Optional[Union[int, str]] = None
    sheet: Optional[str] = None


@dataclass
class FileVerificationResult:
    """Result of file verification"""
    success: bool
    message: str
    file_info: Optional[FileInfo] = None
    csv_data: Optional[CSVData] = None
    excel_data: Optional[ExcelData] = None
    json_data: Optional[Any] = None
    image_info: Optional[ImageInfo] = None
    extracted_values: Dict[str, Any] = field(default_factory=dict)
    assertion_results: List[Dict[str, Any]] = field(default_factory=list)
    duration_ms: int = 0


class FileVerificationService:
    """
    Service for file download verification in automated tests.
    """
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check available file parsing libraries"""
        self.has_pandas = False
        self.has_openpyxl = False
        self.has_pillow = False
        self.has_xmltodict = False
        
        try:
            import pandas
            self.has_pandas = True
        except ImportError:
            logger.debug("pandas not installed (optional for CSV). Run: pip install pandas")
        
        try:
            import openpyxl
            self.has_openpyxl = True
        except ImportError:
            logger.debug("openpyxl not installed (optional for Excel). Run: pip install openpyxl")
        
        try:
            from PIL import Image
            self.has_pillow = True
        except ImportError:
            logger.debug("Pillow not installed (optional for images). Run: pip install Pillow")
        
        try:
            import xmltodict
            self.has_xmltodict = True
        except ImportError:
            logger.debug("xmltodict not installed (optional for XML). Run: pip install xmltodict")
    
    def get_file_info(self, file_path: str) -> Optional[FileInfo]:
        """Get basic information about a file"""
        if not os.path.exists(file_path):
            return None
        
        path = Path(file_path)
        stat = path.stat()
        
        # Try to determine MIME type
        mime_type = None
        try:
            import mimetypes
            mime_type, _ = mimetypes.guess_type(file_path)
        except:
            pass
        
        return FileInfo(
            path=str(path.absolute()),
            name=path.name,
            extension=path.suffix.lower(),
            size_bytes=stat.st_size,
            mime_type=mime_type,
            created_at=datetime.fromtimestamp(stat.st_ctime),
            modified_at=datetime.fromtimestamp(stat.st_mtime)
        )
    
    def parse_csv(self, file_path: str, delimiter: str = ',', encoding: str = 'utf-8') -> Optional[CSVData]:
        """Parse a CSV file"""
        try:
            if self.has_pandas:
                import pandas as pd
                df = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)
                return CSVData(
                    headers=list(df.columns),
                    rows=[list(row) for row in df.values],
                    row_count=len(df),
                    column_count=len(df.columns)
                )
            else:
                # Fallback to csv module
                import csv
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = list(reader)
                    if not rows:
                        return CSVData(headers=[], rows=[], row_count=0, column_count=0)
                    headers = rows[0]
                    data_rows = rows[1:]
                    return CSVData(
                        headers=headers,
                        rows=data_rows,
                        row_count=len(data_rows),
                        column_count=len(headers)
                    )
        except Exception as e:
            logger.error(f"Failed to parse CSV: {e}")
            return None
    
    def parse_excel(self, file_path: str) -> Optional[ExcelData]:
        """Parse an Excel file"""
        if not self.has_openpyxl:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return None
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                headers = []
                
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if i == 0:
                        headers = row_data
                    else:
                        rows.append(row_data)
                
                sheets[sheet_name] = CSVData(
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    column_count=len(headers)
                )
            
            return ExcelData(sheets=sheets, sheet_names=wb.sheetnames)
            
        except Exception as e:
            logger.error(f"Failed to parse Excel: {e}")
            return None
    
    def parse_json(self, file_path: str) -> Optional[Any]:
        """Parse a JSON file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to parse JSON: {e}")
            return None
    
    def parse_xml(self, file_path: str) -> Optional[Dict]:
        """Parse an XML file"""
        if not self.has_xmltodict:
            logger.error("xmltodict not installed. Run: pip install xmltodict")
            return None
        
        try:
            import xmltodict
            with open(file_path, 'r', encoding='utf-8') as f:
                return xmltodict.parse(f.read())
        except Exception as e:
            logger.error(f"Failed to parse XML: {e}")
            return None
    
    def get_image_info(self, file_path: str) -> Optional[ImageInfo]:
        """Get image file information"""
        if not self.has_pillow:
            logger.error("Pillow not installed. Run: pip install Pillow")
            return None
        
        try:
            from PIL import Image
            with Image.open(file_path) as img:
                return ImageInfo(
                    width=img.width,
                    height=img.height,
                    format=img.format or 'unknown',
                    mode=img.mode,
                    has_transparency=img.mode in ('RGBA', 'LA', 'P')
                )
        except Exception as e:
            logger.error(f"Failed to get image info: {e}")
            return None
    
    def get_json_path(self, data: Any, path: str) -> Any:
        """
        Get value from JSON data using JSONPath-like syntax.
        Supports: $.key.subkey, $.array[0], $.array[*].field
        """
        if path.startswith('$.'):
            path = path[2:]
        
        parts = re.split(r'\.|\[|\]', path)
        parts = [p for p in parts if p]  # Remove empty parts
        
        current = data
        for part in parts:
            if current is None:
                return None
            
            if part == '*':
                # Wildcard - return all items
                if isinstance(current, list):
                    return current
                return None
            
            if part.isdigit():
                # Array index
                idx = int(part)
                if isinstance(current, list) and idx < len(current):
                    current = current[idx]
                else:
                    return None
            else:
                # Object key
                if isinstance(current, dict):
                    current = current.get(part)
                else:
                    return None
        
        return current
    
    async def verify_file(
        self,
        file_path: str,
        file_type: str = "auto",  # "auto", "csv", "excel", "json", "xml", "image", "any"
        assertions: List[FileAssertion] = None,
        csv_options: Optional[Dict[str, str]] = None,  # delimiter, encoding
        extract_value: Optional[Dict[str, Any]] = None  # {"path": "...", "store_as": "..."}
    ) -> FileVerificationResult:
        """
        Verify a downloaded file.
        
        Args:
            file_path: Path to the file
            file_type: Type of file or "auto" to detect
            assertions: List of assertions to verify
            csv_options: Options for CSV parsing
            extract_value: Extract value from file
            
        Returns:
            FileVerificationResult with parsed data and assertion results
        """
        start_time = datetime.now()
        assertions = assertions or []
        csv_options = csv_options or {}
        extracted_values = {}
        assertion_results = []
        
        # Get file info
        file_info = self.get_file_info(file_path)
        if not file_info:
            return FileVerificationResult(
                success=False,
                message=f"File not found: {file_path}",
                duration_ms=int((datetime.now() - start_time).total_seconds() * 1000)
            )
        
        # Detect file type if auto
        if file_type == "auto":
            ext = file_info.extension
            if ext in ['.csv', '.tsv']:
                file_type = "csv"
            elif ext in ['.xlsx', '.xls']:
                file_type = "excel"
            elif ext == '.json':
                file_type = "json"
            elif ext == '.xml':
                file_type = "xml"
            elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
                file_type = "image"
            else:
                file_type = "any"
        
        # Parse file based on type
        csv_data = None
        excel_data = None
        json_data = None
        image_info = None
        
        if file_type == "csv":
            csv_data = self.parse_csv(
                file_path,
                delimiter=csv_options.get('delimiter', ','),
                encoding=csv_options.get('encoding', 'utf-8')
            )
        elif file_type == "excel":
            excel_data = self.parse_excel(file_path)
        elif file_type == "json":
            json_data = self.parse_json(file_path)
        elif file_type == "xml":
            json_data = self.parse_xml(file_path)  # Convert to dict
        elif file_type == "image":
            image_info = self.get_image_info(file_path)
        
        # Run assertions
        all_passed = True
        for assertion in assertions:
            result = self._run_assertion(
                file_info, csv_data, excel_data, json_data, image_info, assertion
            )
            assertion_results.append(result)
            if not result['passed']:
                all_passed = False
        
        # Extract value if requested
        if extract_value:
            path = extract_value.get('path', '')
            store_as = extract_value.get('store_as', 'extractedValue')
            row = extract_value.get('row')
            col = extract_value.get('col')
            sheet = extract_value.get('sheet')
            
            value = None
            
            if json_data is not None:
                value = self.get_json_path(json_data, path)
            elif csv_data and row is not None:
                value = csv_data.get_cell(row, col or 0)
            elif excel_data and sheet:
                sheet_data = excel_data.sheets.get(sheet)
                if sheet_data and row is not None:
                    value = sheet_data.get_cell(row, col or 0)
            
            if value is not None:
                extracted_values[store_as] = value
            else:
                all_passed = False
                assertion_results.append({
                    'type': 'extract_value',
                    'passed': False,
                    'message': f"Could not extract value at path: {path}"
                })
        
        duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        
        return FileVerificationResult(
            success=all_passed,
            message="All file assertions passed" if all_passed else "Some assertions failed",
            file_info=file_info,
            csv_data=csv_data,
            excel_data=excel_data,
            json_data=json_data,
            image_info=image_info,
            extracted_values=extracted_values,
            assertion_results=assertion_results,
            duration_ms=duration_ms
        )
    
    def _run_assertion(
        self,
        file_info: FileInfo,
        csv_data: Optional[CSVData],
        excel_data: Optional[ExcelData],
        json_data: Optional[Any],
        image_info: Optional[ImageInfo],
        assertion: FileAssertion
    ) -> Dict[str, Any]:
        """Run a single assertion"""
        assertion_type = assertion.type
        expected = assertion.expected
        
        passed = False
        message = ""
        actual = ""
        
        try:
            # File info assertions
            if assertion_type == 'file_exists':
                passed = file_info is not None
                message = f"File {'exists' if passed else 'does not exist'}"
                
            elif assertion_type == 'file_name_contains':
                actual = file_info.name
                passed = expected.lower() in actual.lower()
                message = f"File name '{actual}' {'contains' if passed else 'does not contain'} '{expected}'"
                
            elif assertion_type == 'file_name_equals':
                actual = file_info.name
                passed = actual == expected
                message = f"File name '{actual}' {'equals' if passed else 'does not equal'} '{expected}'"
                
            elif assertion_type == 'file_extension':
                actual = file_info.extension
                passed = actual.lower() == expected.lower() or actual.lower() == f".{expected.lower()}"
                message = f"File extension '{actual}' {'is' if passed else 'is not'} '{expected}'"
                
            elif assertion_type == 'size_min':
                actual = str(file_info.size_bytes)
                passed = file_info.size_bytes >= int(expected)
                message = f"File size {actual} bytes {'≥' if passed else '<'} {expected}"
                
            elif assertion_type == 'size_max':
                actual = str(file_info.size_bytes)
                passed = file_info.size_bytes <= int(expected)
                message = f"File size {actual} bytes {'≤' if passed else '>'} {expected}"
                
            elif assertion_type == 'size_equals':
                actual = str(file_info.size_bytes)
                passed = file_info.size_bytes == int(expected)
                message = f"File size {actual} bytes {'equals' if passed else 'does not equal'} {expected}"
                
            # CSV assertions
            elif assertion_type == 'csv_row_count':
                if csv_data:
                    actual = str(csv_data.row_count)
                    passed = csv_data.row_count == int(expected)
                    message = f"CSV has {actual} rows, expected {expected}"
                else:
                    message = "CSV data not available"
                    
            elif assertion_type == 'csv_row_count_min':
                if csv_data:
                    actual = str(csv_data.row_count)
                    passed = csv_data.row_count >= int(expected)
                    message = f"CSV has {actual} rows {'≥' if passed else '<'} {expected}"
                else:
                    message = "CSV data not available"
                    
            elif assertion_type == 'csv_column_count':
                if csv_data:
                    actual = str(csv_data.column_count)
                    passed = csv_data.column_count == int(expected)
                    message = f"CSV has {actual} columns, expected {expected}"
                else:
                    message = "CSV data not available"
                    
            elif assertion_type == 'csv_header_contains':
                if csv_data:
                    passed = expected in csv_data.headers
                    message = f"CSV headers {'contain' if passed else 'do not contain'} '{expected}'"
                else:
                    message = "CSV data not available"
                    
            elif assertion_type == 'csv_cell_equals':
                if csv_data:
                    row = assertion.row or 0
                    col = assertion.col or 0
                    actual = csv_data.get_cell(row, col) or ""
                    passed = actual == expected
                    message = f"CSV cell [{row},{col}] is '{actual}', expected '{expected}'"
                else:
                    message = "CSV data not available"
                    
            elif assertion_type == 'csv_cell_contains':
                if csv_data:
                    row = assertion.row or 0
                    col = assertion.col or 0
                    actual = csv_data.get_cell(row, col) or ""
                    passed = expected.lower() in actual.lower()
                    message = f"CSV cell [{row},{col}] '{actual}' {'contains' if passed else 'does not contain'} '{expected}'"
                else:
                    message = "CSV data not available"
                    
            # Excel assertions
            elif assertion_type == 'excel_sheet_exists':
                if excel_data:
                    passed = expected in excel_data.sheet_names
                    message = f"Excel {'has' if passed else 'does not have'} sheet '{expected}'"
                else:
                    message = "Excel data not available"
                    
            elif assertion_type == 'excel_sheet_count':
                if excel_data:
                    actual = str(len(excel_data.sheet_names))
                    passed = len(excel_data.sheet_names) == int(expected)
                    message = f"Excel has {actual} sheets, expected {expected}"
                else:
                    message = "Excel data not available"
                    
            # JSON assertions
            elif assertion_type == 'json_path_equals':
                if json_data is not None:
                    path = assertion.col or '$'  # Use col field for path
                    actual = str(self.get_json_path(json_data, str(path)))
                    passed = actual == expected
                    message = f"JSON path '{path}' is '{actual}', expected '{expected}'"
                else:
                    message = "JSON data not available"
                    
            elif assertion_type == 'json_path_exists':
                if json_data is not None:
                    path = expected
                    value = self.get_json_path(json_data, path)
                    passed = value is not None
                    message = f"JSON path '{path}' {'exists' if passed else 'does not exist'}"
                else:
                    message = "JSON data not available"
                    
            elif assertion_type == 'json_array_length':
                if json_data is not None:
                    path = assertion.col or '$'
                    arr = self.get_json_path(json_data, str(path))
                    if isinstance(arr, list):
                        actual = str(len(arr))
                        passed = len(arr) == int(expected)
                        message = f"JSON array at '{path}' has {actual} items, expected {expected}"
                    else:
                        message = f"JSON path '{path}' is not an array"
                else:
                    message = "JSON data not available"
                    
            # Image assertions
            elif assertion_type == 'image_width':
                if image_info:
                    actual = str(image_info.width)
                    passed = image_info.width == int(expected)
                    message = f"Image width is {actual}px, expected {expected}px"
                else:
                    message = "Image data not available"
                    
            elif assertion_type == 'image_height':
                if image_info:
                    actual = str(image_info.height)
                    passed = image_info.height == int(expected)
                    message = f"Image height is {actual}px, expected {expected}px"
                else:
                    message = "Image data not available"
                    
            elif assertion_type == 'image_format':
                if image_info:
                    actual = image_info.format
                    passed = actual.lower() == expected.lower()
                    message = f"Image format is '{actual}', expected '{expected}'"
                else:
                    message = "Image data not available"
                    
            elif assertion_type == 'image_min_width':
                if image_info:
                    actual = str(image_info.width)
                    passed = image_info.width >= int(expected)
                    message = f"Image width {actual}px {'≥' if passed else '<'} {expected}px"
                else:
                    message = "Image data not available"
                    
            elif assertion_type == 'image_min_height':
                if image_info:
                    actual = str(image_info.height)
                    passed = image_info.height >= int(expected)
                    message = f"Image height {actual}px {'≥' if passed else '<'} {expected}px"
                else:
                    message = "Image data not available"
                    
            else:
                message = f"Unknown assertion type: {assertion_type}"
                
        except Exception as e:
            message = f"Assertion error: {str(e)}"
        
        return {
            'type': assertion_type,
            'expected': expected,
            'actual': actual,
            'passed': passed,
            'message': message
        }


# Singleton instance
_file_service: Optional[FileVerificationService] = None

def get_file_service() -> FileVerificationService:
    """Get or create the file verification service"""
    global _file_service
    if _file_service is None:
        _file_service = FileVerificationService()
    return _file_service

